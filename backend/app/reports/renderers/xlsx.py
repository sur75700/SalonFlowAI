from __future__ import annotations

from datetime import UTC
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.writer.excel import ExcelWriter

from app.reports.models import DailySummaryReport
from app.reports.renderers import (
    format_report_datetime,
    report_status,
    report_text,
    spreadsheet_safe,
)


def render_daily_summary_xlsx(report: DailySummaryReport) -> bytes:
    locale = report.locale
    metrics = report.metrics

    workbook = Workbook()
    package_timestamp = (
        report.generated_at.astimezone(UTC).replace(tzinfo=None)
    )
    workbook.properties.created = package_timestamp
    workbook.properties.modified = package_timestamp

    overview = workbook.active
    overview.title = "Overview"

    overview_rows = [
        [report_text(locale, "metric"), report_text(locale, "value")],
        [report_text(locale, "report_date"), report.report_date.isoformat()],
        [report_text(locale, "generated_at"), report.generated_at.strftime("%Y-%m-%d %H:%M UTC")],
        [report_text(locale, "total_clients"), metrics.total_clients],
        [report_text(locale, "total_services"), metrics.total_services],
        [report_text(locale, "total_appointments"), metrics.total_appointments],
        [report_text(locale, "appointments_on_date"), metrics.appointments_on_date],
        [report_text(locale, "scheduled_on_date"), metrics.scheduled_on_date],
        [report_text(locale, "completed_on_date"), metrics.completed_on_date],
        [report_text(locale, "cancelled_on_date"), metrics.cancelled_on_date],
    ]

    for row in overview_rows:
        overview.append(row)

    for cell in overview[1]:
        cell.font = Font(bold=True)

    overview.freeze_panes = "A2"
    overview.column_dimensions["A"].width = 38
    overview.column_dimensions["B"].width = 28

    appointments = workbook.create_sheet("Appointments")
    appointments.append(
        [
            report_text(locale, "start"),
            report_text(locale, "client"),
            report_text(locale, "service"),
            report_text(locale, "status"),
            report_text(locale, "notes"),
        ]
    )

    for cell in appointments[1]:
        cell.font = Font(bold=True)

    for item in report.appointments:
        appointments.append(
            [
                spreadsheet_safe(format_report_datetime(item.starts_at)),
                spreadsheet_safe(item.client_name),
                spreadsheet_safe(item.service_name),
                spreadsheet_safe(report_status(locale, item.status)),
                spreadsheet_safe(item.notes),
            ]
        )

    appointments.freeze_panes = "A2"
    appointments.column_dimensions["A"].width = 22
    appointments.column_dimensions["B"].width = 28
    appointments.column_dimensions["C"].width = 28
    appointments.column_dimensions["D"].width = 18
    appointments.column_dimensions["E"].width = 48

    buffer = BytesIO()
    archive = ZipFile(
        buffer,
        "w",
        ZIP_DEFLATED,
        allowZip64=True,
    )
    ExcelWriter(workbook, archive).save()
    workbook.close()
    return buffer.getvalue()

# PHASE_63D_REPORT_DOCUMENT_RENDERER
def render_report_document_xlsx(document: object) -> bytes:
    from datetime import UTC
    from io import BytesIO
    from zipfile import ZIP_DEFLATED, ZipFile

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.writer.excel import ExcelWriter

    from app.reports.contracts import ReportDocument

    if not isinstance(document, ReportDocument):
        raise TypeError("document must be a ReportDocument")

    def safe(value: object) -> object:
        if not isinstance(value, str):
            return value
        return f"'{value}" if value.startswith(("=", "+", "-", "@")) else value

    workbook = Workbook()
    package_timestamp = (
        document.generated_at.astimezone(UTC).replace(tzinfo=None)
    )
    workbook.properties.created = package_timestamp
    workbook.properties.modified = package_timestamp

    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Report Type", safe(document.report_type)])
    overview.append(["Start Date", document.period.start_date.isoformat()])
    overview.append(["End Date", document.period.end_date.isoformat()])
    overview.append(["Timezone", safe(document.period.timezone)])
    overview.append(["Locale", safe(document.locale)])
    overview.append([])
    overview.append(["Metric", "Value"])
    for key, value in document.metrics.items():
        overview.append([safe(str(key)), safe(value)])
    for cell in overview[7]:
        cell.font = Font(bold=True)

    rows_sheet = workbook.create_sheet("Rows")
    if document.columns:
        rows_sheet.append([safe(value) for value in document.columns])
        for cell in rows_sheet[1]:
            cell.font = Font(bold=True)
        for row in document.rows:
            rows_sheet.append([safe(value) for value in row])

    buffer = BytesIO()
    archive = ZipFile(
        buffer,
        "w",
        ZIP_DEFLATED,
        allowZip64=True,
    )
    ExcelWriter(workbook, archive).save()
    workbook.close()
    return buffer.getvalue()
